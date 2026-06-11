"""
Lab 5 Team Work Allocation - Clean Excel Generator
"""
import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ── Colours ───────────────────────────────────────────────────────────────────
NAVY      = "1B2A4A"
BLUE      = "2E5EAA"
SKY       = "5B9BD5"
PALE      = "EDF3FA"
STRIPE    = "F5F8FC"
WHITE     = "FFFFFF"
TEXT      = "2D3436"
TEXT_MID  = "636E72"
BORDER_C  = "BFCFE0"
GREEN_BG  = "D5F5E3"
GRAY_BG   = "F0F0F0"
RED_BG    = "FADBD8"

FONT_NAME = "Calibri"
SC = 2          # first data column (B). Column A = left margin.
EC = 10         # last data column (J). Column K = right margin.

wb = Workbook()

# ── Style helpers ─────────────────────────────────────────────────────────────
def mk_font(bold=False, size=11, color=TEXT, italic=False):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color, italic=italic)

def mk_fill(color):
    return PatternFill("solid", fgColor=color)

def mk_border():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def mk_align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cell(ws, r, c, val=None, bold=False, size=11, color=TEXT, bg=None,
         h="left", v="center", wrap=True, border=True, italic=False):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = mk_font(bold, size, color, italic)
    if bg:
        cl.fill = mk_fill(bg)
    cl.alignment = mk_align(h, v, wrap)
    if border:
        cl.border = mk_border()
    return cl

def row_h(ws, r, h):
    ws.row_dimensions[r].height = h

def col_w(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

def setup_ws(ws, tab_color=BLUE, zoom=100):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    ws.sheet_properties.tabColor = tab_color
    ws.page_margins = PageMargins(left=0.6, right=0.6, top=0.7, bottom=0.7)
    col_w(ws, {"A": 2, "K": 2})   # side margins

def paint_range(ws, r1, c1, r2, c2, bg):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = mk_fill(bg)

def banner(ws, r, title, subtitle):
    ws.merge_cells(start_row=r, start_column=SC, end_row=r, end_column=EC)
    cell(ws, r, SC, title, bold=True, size=15, color=WHITE, bg=NAVY, h="center")
    paint_range(ws, r, SC, r, EC, NAVY)
    row_h(ws, r, 34)

    ws.merge_cells(start_row=r + 1, start_column=SC, end_row=r + 1, end_column=EC)
    cell(ws, r + 1, SC, subtitle, size=10, color="C8D6E8", bg=NAVY, h="center")
    paint_range(ws, r + 1, SC, r + 1, EC, NAVY)
    row_h(ws, r + 1, 20)
    return r + 3

def section(ws, r, text):
    ws.merge_cells(start_row=r, start_column=SC, end_row=r, end_column=EC)
    cell(ws, r, SC, text, bold=True, size=11, color=WHITE, bg=BLUE, h="left")
    paint_range(ws, r, SC, r, EC, BLUE)
    row_h(ws, r, 24)
    return r + 1

def info_grid(ws, r, items):
    """Two-column key-value pairs in a clean grid (label | value | label | value)."""
    pairs_per_row = 2
    for i in range(0, len(items), pairs_per_row):
        chunk = items[i:i + pairs_per_row]
        row_h(ws, r, 22)
        col_positions = [(SC, SC + 1), (SC + 3, SC + 4)]
        for j, (k, v) in enumerate(chunk):
            lc, vc = col_positions[j]
            cell(ws, r, lc, k, bold=True, size=10, color=NAVY, bg=PALE, h="left")
            ws.merge_cells(start_row=r, start_column=vc, end_row=r, end_column=vc + 1)
            cell(ws, r, vc, v, size=10, bg=WHITE, h="left")
            paint_range(ws, r, vc, r, vc + 1, WHITE)
        r += 1
    return r + 1

def thead(ws, r, headers, heights=28):
    for i, h in enumerate(headers):
        cell(ws, r, SC + i, h, bold=True, size=10, color=WHITE, bg=BLUE, h="center")
    row_h(ws, r, heights)
    return r + 1

def calc_row_height(values, base=22, per_line=14):
    lines = 1
    for v in values:
        if v is None:
            continue
        lines = max(lines, str(v).count("\n") + 1)
    return max(base, lines * per_line + 8)

def trow(ws, r, values, stripe=False, centers=None, bolds=None, heights=None):
    bg = STRIPE if stripe else WHITE
    centers = centers or set()
    bolds = bolds or set()
    h = heights or calc_row_height(values)
    row_h(ws, r, h)
    for i, val in enumerate(values):
        c = SC + i
        cell(ws, r, c, val,
             bold=(i in bolds), size=10,
             color=BLUE if i in bolds else TEXT,
             bg=bg,
             h="center" if i in centers else "left")
    return r + 1

def tfoot(ws, r, label, nums, merge_count):
    ws.merge_cells(start_row=r, start_column=SC, end_row=r, end_column=SC + merge_count - 1)
    cell(ws, r, SC, label, bold=True, size=11, color=WHITE, bg=SKY, h="right")
    paint_range(ws, r, SC, r, SC + merge_count - 1, SKY)
    for i, n in enumerate(nums):
        cell(ws, r, SC + merge_count + i, n, bold=True, size=11, color=WHITE, bg=SKY, h="center")
    row_h(ws, r, 26)
    return r + 1

# ═══════════════════════════════════════════════════════════════════════════════
# DATA
# ═══════════════════════════════════════════════════════════════════════════════
team = [
    (1,  "Hassana", "Schema & Failed Queue Developer",       "Task 1, Task 10",
     "initialize_productionops_schema.sql\nfailed_inspection_queue.sql", 11, 7, "TBD"),
    (2,  "Shrili",  "Inspection & Temp Batch Developer",     "Task 2, Task 6",
     "inspection_request_registration.sql\ninspection_batch_processing.sql", 11, 7, "TBD"),
    (3,  "Brain",   "Control Flow Developer",                "Task 3, Task 4",
     "inspection_result_classification.sql\nmonthly_inspection_schedule.sql", 11, 7, "TBD"),
    (4,  "Parth",   "Error Handling Developer",              "Task 5",
     "error_log.sql\ninspection_error_logging.sql", 11, 7, "TBD"),
    (5,  "Josovo",  "Reporting & Views Developer",           "Task 8, Task 9",
     "product_quality_statistics.sql\nvProductInspectionSummary.sql", 11, 7, "TBD"),
    (6,  "Lien",    "Release Review Table Developer",        "Task 12",
     "product_release_review.sql", 11, 7, "TBD"),
    (7,  "Kelvin",  "Categories & Notification Developer",   "Task 7, Task 13",
     "inspection_category_management.sql\nnotification_log.sql\ninspection_notification_cursor.sql", 11, 7, "TBD"),
    (8,  "Dhruv",   "Failed Review Cursor Developer",        "Task 11",
     "failed_product_review_cursor.sql", 11, 7, "TBD"),
    (9,  "Sahil",   "Deployment & Final Submission (LAST)",  "Task 14 + Final Submission",
     "deploy_all.sql\nvalidate_lab5_results.sql\nREADME.md\nData Mapping\nScreenshots (15)\nQA Review", 11, 7, "TBD"),
]

scripts = [
    ("Task 1",  "initialize_productionops_schema.sql",        "scripts/setup/",             "Hassana",      "Schema",           11, 7),
    ("Task 10", "failed_inspection_queue.sql",                "scripts/tables/",            "Hassana",      "Table",            11, 7),
    ("Task 2",  "inspection_request_registration.sql",        "scripts/tables/",            "Shrili",       "Table",            11, 7),
    ("Task 6",  "inspection_batch_processing.sql",            "scripts/temporary_objects/", "Shrili",       "Temp Table",       11, 7),
    ("Task 3",  "inspection_result_classification.sql",       "scripts/controlflow/",       "Brain",        "Control Flow",     11, 7),
    ("Task 4",  "monthly_inspection_schedule.sql",            "scripts/controlflow/",       "Brain",        "Control Flow",     11, 7),
    ("Task 5",  "error_log.sql",                              "scripts/tables/",            "Parth",        "Table",            11, 7),
    ("Task 5",  "inspection_error_logging.sql",               "scripts/errorhandling/",     "Parth",        "Error Handling",   11, 7),
    ("Task 8",  "product_quality_statistics.sql",             "scripts/reporting/",         "Josovo",       "Derived Table",    11, 7),
    ("Task 9",  "vProductInspectionSummary.sql",              "scripts/reporting/",         "Josovo",       "View",             11, 7),
    ("Task 12", "product_release_review.sql",                 "scripts/tables/",            "Lien",         "Table",            11, 7),
    ("Task 7",  "inspection_category_management.sql",         "scripts/temporary_objects/", "Kelvin",       "Table Variable",   11, 7),
    ("Task 13", "notification_log.sql",                       "scripts/tables/",            "Kelvin",       "Table",            11, 7),
    ("Task 13", "inspection_notification_cursor.sql",         "scripts/cursors/",           "Kelvin",       "Cursor",           11, 7),
    ("Task 11", "failed_product_review_cursor.sql",           "scripts/cursors/",           "Dhruv",        "Cursor",           11, 7),
    ("Task 14", "deploy_all.sql",                             "scripts/deployment/",        "Sahil (LAST)", "Deployment",       11, 7),
    ("Task 14", "validate_lab5_results.sql",                  "scripts/validation/",        "Sahil (LAST)", "Validation",       11, 7),
    ("Final",   "README.md + Data Mapping + Screenshots + QA",  "Lab5/",                      "Sahil (LAST)", "Final Submission", 11, 7),
]

deps = [
    ("Task 1 - Schema",              "Hassana",      "None",            "All members",           "Hard Blocker", "CRITICAL - stops team"),
    ("Task 10 - Failed Queue",       "Hassana",      "Task 1",          "Task 11 (Dhruv)",       "Hard Blocker", "Dhruv blocked Day 4"),
    ("Task 2 - Inspection Requests", "Shrili",       "Task 1",          "Validation",            "Soft",         "Low risk"),
    ("Task 6 - Temp Batch",          "Shrili",       "Task 1",          "Task 14 (Sahil)",       "Soft",         "Independent"),
    ("Task 3 - IF/ELSE",             "Brain",        "Task 1",          "Task 14 (Sahil)",       "Soft",         "Parallel Day 2"),
    ("Task 4 - WHILE",               "Brain",        "Task 1",          "Task 14 (Sahil)",       "Soft",         "Parallel Day 2"),
    ("Task 5 - Error Logging",       "Parth",        "Task 1",          "Task 14 (Sahil)",       "Hard Blocker", "ErrorLog required"),
    ("Task 8 - Derived Tables",      "Josovo",       "Task 1",          "Task 14 (Sahil)",       "Soft",         "Independent"),
    ("Task 9 - View",                "Josovo",       "Task 1",          "Task 14 (Sahil)",       "Soft",         "Independent"),
    ("Task 12 - Release Review",     "Lien",         "Task 1",          "Task 11 (Dhruv)",       "Hard Blocker", "Cursor insert target"),
    ("Task 7 - Table Variables",     "Kelvin",       "Task 1",          "Task 14 (Sahil)",       "Soft",         "Independent"),
    ("Task 11 - Review Cursor",      "Dhruv",        "Task 10 + 12",    "Task 13 (Kelvin)",      "Hard Blocker", "Needs queue + table"),
    ("Task 13 - Notification",       "Kelvin",       "Task 11",         "Task 14 (Sahil)",       "Hard Blocker", "After review cursor"),
    ("Task 14 - Final Submission",   "Sahil (LAST)", "Tasks 1-13",      "GitHub submission",     "Hard Blocker", "Cannot submit early"),
]

start_date = date(2026, 6, 10)
timeline = [
    ("Day 1", start_date,              "Foundation",    "Hassana",      "Task 1 - Schema",                    "Schema tested",              11, 7),
    ("Day 1", start_date,              "Foundation",    "Shrili",       "Task 2 - InspectionRequests",          "10 inserts done",            11, 7),
    ("Day 1", start_date,              "Foundation",    "Lien",         "Task 12 - ProductReleaseReview",       "Table DDL done",             11, 7),
    ("Day 2", start_date + timedelta(1), "Development", "Hassana",      "Task 10 - FailedInspectionQueue",      "Queue populated",            11, 7),
    ("Day 2", start_date + timedelta(1), "Development", "Shrili",       "Task 6 - Temp Batch",                  "Batch script done",          11, 7),
    ("Day 2", start_date + timedelta(1), "Development", "Brain",        "Task 3 + 4 - IF/ELSE + WHILE",         "Control flow done",          11, 7),
    ("Day 2", start_date + timedelta(1), "Development", "Parth",        "Task 5 - Error Logging",               "TRY/CATCH working",          11, 7),
    ("Day 2", start_date + timedelta(1), "Development", "Josovo",       "Task 8 + 9 - Derived + View",          "Reporting done",             11, 7),
    ("Day 2", start_date + timedelta(1), "Development", "Kelvin",       "Task 7 - Table Variables",             "Categories done",            11, 7),
    ("Day 3", start_date + timedelta(2), "Testing",     "All except Sahil", "Test own scripts in SSMS",         "No errors",                  11, 7),
    ("Day 4", start_date + timedelta(3), "Cursors",     "Dhruv",        "Task 11 - Review Cursor",              "All Pending processed",      11, 7),
    ("Day 4", start_date + timedelta(3), "Cursors",     "Kelvin",       "Task 13 - Notification Cursor",        "Notifications inserted",     11, 7),
    ("Day 5", start_date + timedelta(4), "Handoff",     "All",          "Push all scripts to GitHub",           "Repo complete",              11, 7),
    ("Day 6", start_date + timedelta(5), "Final Task",  "Sahil (LAST)", "Task 14 - deploy + validate + README", "Deployment green",           11, 7),
    ("Day 7", start_date + timedelta(6), "Submission",  "Sahil (LAST)", "Screenshots + QA + submit link",       "Submitted to professor",     11, 7),
]

sp_data = [
    ("Hassana",      "Task 1 + Task 10",             "Medium", 11, 7, "Day 1-2",  "Schema + Failed Queue"),
    ("Shrili",       "Task 2 + Task 6",              "Medium", 11, 7, "Day 1-2",  "Tables + Temp Batch"),
    ("Brain",        "Task 3 + Task 4",             "Medium", 11, 7, "Day 2",    "IF/ELSE + WHILE"),
    ("Parth",        "Task 5",                       "High",   11, 7, "Day 2",    "Error Handling"),
    ("Josovo",       "Task 8 + Task 9",              "Medium", 11, 7, "Day 2",    "Derived + View"),
    ("Lien",         "Task 12",                      "Medium", 11, 7, "Day 1",    "Release Review Table"),
    ("Kelvin",       "Task 7 + Task 13",             "Medium", 11, 7, "Day 2-4",  "Categories + Notification"),
    ("Dhruv",        "Task 11",                      "High",   11, 7, "Day 4",    "Review Cursor"),
    ("Sahil (LAST)", "Task 14 + Final Submission",   "High",   11, 7, "Day 6-7",  "Deploy + Docs + QA"),
]

screenshots = [
    (1,  "ProductionOps schema validation",      "Hassana", "initialize_productionops_schema.sql",  "After Task 1",  "01_schema_validation.png"),
    (2,  "FailedInspectionQueue data",           "Hassana", "failed_inspection_queue.sql",          "After Task 10", "02_failed_queue.png"),
    (3,  "InspectionRequests data",              "Shrili",  "inspection_request_registration.sql",  "After Task 2",  "03_inspection_requests.png"),
    (4,  "#InspectionBatch temp table",          "Shrili",  "inspection_batch_processing.sql",      "After Task 6",  "04_inspection_batch.png"),
    (5,  "IF/ELSE classification results",       "Brain",   "inspection_result_classification.sql", "After Task 3",  "05_if_else.png"),
    (6,  "WHILE monthly schedule",               "Brain",   "monthly_inspection_schedule.sql",      "After Task 4",  "06_while_schedule.png"),
    (7,  "ErrorLog contents",                    "Parth",   "inspection_error_logging.sql",         "After Task 5",  "07_error_log.png"),
    (8,  "Derived table statistics",             "Josovo",  "product_quality_statistics.sql",       "After Task 8",  "08_quality_stats.png"),
    (9,  "vProductInspectionSummary view",       "Josovo",  "vProductInspectionSummary.sql",        "After Task 9",  "09_view_output.png"),
    (10, "ProductReleaseReview table",           "Lien",    "product_release_review.sql",           "After Task 12", "10_release_review.png"),
    (11, "@InspectionCategories output",         "Kelvin",  "inspection_category_management.sql",   "After Task 7",  "11_categories.png"),
    (12, "NotificationLog contents",             "Kelvin",  "inspection_notification_cursor.sql",   "After Task 13", "12_notification_log.png"),
    (13, "Review cursor execution",              "Dhruv",   "failed_product_review_cursor.sql",     "During Task 11","13_review_cursor.png"),
    (14, "ProductReleaseReview after cursor",    "Dhruv",   "failed_product_review_cursor.sql",     "After Task 11", "14_review_results.png"),
    (15, "deploy_all.sql success",               "Sahil",   "deploy_all.sql",                       "After Task 14", "15_deploy_success.png"),
]

quick_ref = [
    ("Hassana",      "Task 1, Task 10",            "initialize_productionops_schema.sql\nfailed_inspection_queue.sql",
     "None - start Day 1",              "2 screenshots",              "N/A"),
    ("Shrili",       "Task 2, Task 6",             "inspection_request_registration.sql\ninspection_batch_processing.sql",
     "Task 1 (Hassana)",                "2 screenshots",              "Hassana"),
    ("Brain",        "Task 3, Task 4",             "inspection_result_classification.sql\nmonthly_inspection_schedule.sql",
     "Task 1 (Hassana)",                "2 screenshots",              "Hassana"),
    ("Parth",        "Task 5",                     "error_log.sql\ninspection_error_logging.sql",
     "Task 1 (Hassana)",                "1 screenshot",               "Hassana"),
    ("Josovo",       "Task 8, Task 9",             "product_quality_statistics.sql\nvProductInspectionSummary.sql",
     "Task 1 (Hassana)",                "2 screenshots",              "Hassana"),
    ("Lien",         "Task 12",                    "product_release_review.sql",
     "Task 1 (Hassana)",                "1 screenshot",               "Hassana"),
    ("Kelvin",       "Task 7, Task 13",            "inspection_category_management.sql\nnotification_log.sql\ninspection_notification_cursor.sql",
     "Task 1 + Task 11 (Dhruv)",        "2 screenshots",              "Dhruv"),
    ("Dhruv",        "Task 11",                    "failed_product_review_cursor.sql",
     "Task 10 (Hassana) + Task 12 (Lien)", "2 screenshots",           "Hassana + Lien"),
    ("Sahil (LAST)", "Task 14 + Final Submission", "deploy_all.sql\nvalidate_lab5_results.sql\nREADME + Screenshots + QA",
     "ALL Tasks 1-13 done",             "deploy + compile all 15",  "Any member"),
]

# Standard column widths per sheet type
W_TEAM  = {"B": 5, "C": 14, "D": 22, "E": 16, "F": 34, "G": 10, "H": 10, "I": 14, "J": 4}
W_SCRPT = {"B": 10, "C": 38, "D": 28, "E": 14, "F": 16, "G": 9, "H": 8, "I": 13}
W_DEPS  = {"B": 28, "C": 14, "D": 18, "E": 20, "F": 13, "G": 28}
W_TIME  = {"B": 8, "C": 13, "D": 14, "E": 14, "F": 30, "G": 26, "H": 8, "I": 7, "J": 12}
W_SP    = {"B": 14, "C": 30, "D": 11, "E": 9, "F": 9, "G": 11, "H": 26}
W_SHOT  = {"B": 5, "C": 30, "D": 12, "E": 36, "F": 16, "G": 28, "H": 13}
W_QUICK = {"B": 14, "C": 18, "D": 36, "E": 22, "F": 18, "G": 20}

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 - TEAM OVERVIEW
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Team Overview"
setup_ws(ws1, NAVY, 95)

r = banner(ws1, 1,
           "LAB 5 - TEAM WORK ALLOCATION PLAN",
           "SQL Server Developer  |  T-SQL Control Flow & Data Structures  |  AdventureWorks2022 |  ProductionOps Schema")
row_h(ws1, r, 6); r += 1

r = info_grid(ws1, r, [
    ("Course", "SQL Server Developer"),
    ("Database", "AdventureWorks2022"),
    ("Schema", "ProductionOps"),
    ("Timeline", "7 Days"),
    ("Team Size", "9 Members"),
    ("Workload Each", "11 SP / 7 Hours"),
    ("Total Story Points", "99"),
    ("Total Hours", "63"),
])
row_h(ws1, r, 6); r += 1

r = section(ws1, r, "TEAM MEMBER ASSIGNMENTS  (each task owned by ONE person only)")
hdr = ["#", "Member", "Role", "Lab Task(s)", "Scripts / Deliverables", "SP", "Hrs", "Student No."]
r = thead(ws1, r, hdr, 30)

for i, t in enumerate(team):
    r = trow(ws1, r, list(t), stripe=(i % 2 == 1), centers={0, 5, 6, 7}, bolds={1})

r = tfoot(ws1, r, "TEAM TOTAL  (9 members x 11 SP each)", [99, 63, ""], merge_count=6)
row_h(ws1, r, 8); r += 1

ws1.merge_cells(start_row=r, start_column=SC, end_row=r, end_column=EC)
cell(ws1, r, SC,
     "Sahil owns the LAST task (Task 14 + Final Submission). Replace TBD with student numbers before submitting.",
     size=9, color=TEXT_MID, italic=True, border=False, h="left")
row_h(ws1, r, 18)

col_w(ws1, W_TEAM)
ws1.print_area = f"B1:J{r}"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 - SCRIPT ASSIGNMENT
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Script Assignment")
setup_ws(ws2, BLUE, 95)

r2 = banner(ws2, 1, "SCRIPT ASSIGNMENT MATRIX", "One row per script file - unique task ownership")
row_h(ws2, r2, 6); r2 += 1
r2 = thead(ws2, r2, ["Lab Task", "Script File", "Folder Path", "Owner", "Type", "SP", "Hrs", "Status"])

for i, s in enumerate(scripts):
    row = list(s) + ["Not Started"]
    r2 = trow(ws2, r2, row, stripe=(i % 2 == 1), centers={0, 5, 6, 7})
    cell(ws2, r2 - 1, SC + 7, "Not Started", size=9, color=TEXT_MID, bg=GRAY_BG, h="center")

r2 = tfoot(ws2, r2, "NOTE: Each member = 11 SP total across their scripts", ["", ""], merge_count=6)
col_w(ws2, W_SCRPT)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 - DEPENDENCIES
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Dependencies")
setup_ws(ws3, SKY, 95)

r3 = banner(ws3, 1, "WORK DEPENDENCY MAP", "What must finish before the next task can start")
row_h(ws3, r3, 6); r3 += 1
r3 = thead(ws3, r3, ["Task", "Owner", "Depends On", "Blocks", "Type", "Risk if Delayed"])

for i, d in enumerate(deps):
    r3 = trow(ws3, r3, list(d), stripe=(i % 2 == 1), centers={4})
    if "CRITICAL" in d[5]:
        cell(ws3, r3 - 1, SC + 5, d[5], bold=True, size=10, color="C0392B", bg=RED_BG, h="left")
    elif d[4] == "Hard Blocker":
        cell(ws3, r3 - 1, SC + 4, d[4], bold=True, size=9, color=NAVY, bg=PALE, h="center")

col_w(ws3, W_DEPS)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 - 7-DAY TIMELINE
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("7-Day Timeline")
setup_ws(ws4, "27AE60", 95)

r4 = banner(ws4, 1, "7-DAY SPRINT TIMELINE",
            f"Start: {start_date.strftime('%B %d, %Y')}  |  Equal workload: 11 SP / 7 hrs per member")
row_h(ws4, r4, 6); r4 += 1
r4 = thead(ws4, r4, ["Day", "Date", "Phase", "Member", "Work Item", "Exit Criteria", "SP", "Hrs", "Status"])

for i, t in enumerate(timeline):
    r4 = trow(ws4, r4, list(t) + ["Not Started"], stripe=(i % 2 == 1), centers={0, 1, 6, 7, 8})
    dc = ws4.cell(row=r4 - 1, column=SC + 1, value=t[1])
    dc.number_format = "MMM DD, YYYY"
    dc.alignment = mk_align("center")
    dc.border = mk_border()

col_w(ws4, W_TIME)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 - STORY POINTS
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Story Points")
setup_ws(ws5, "F39C12", 95)

r5 = banner(ws5, 1, "STORY POINT BREAKDOWN", "Equal split: every member = 11 story points / 7 hours")
row_h(ws5, r5, 6); r5 += 1
r5 = thead(ws5, r5, ["Member", "Deliverable", "Complexity", "SP", "Hrs", "Sprint Day", "Marks Covered"])

COMPLEXITY = {"Low": "D5F5E3", "Medium": "D6EAF8", "High": "FADBD8"}
for i, s in enumerate(sp_data):
    r5 = trow(ws5, r5, list(s), stripe=(i % 2 == 1), centers={2, 3, 4, 5}, bolds={0})
    cell(ws5, r5 - 1, SC + 2, s[2], bold=True, size=9, color=NAVY, bg=COMPLEXITY.get(s[2], PALE), h="center")

r5 = tfoot(ws5, r5, "TOTAL", [99, 63, ""], merge_count=4)
col_w(ws5, W_SP)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 - SCREENSHOTS
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Screenshots")
setup_ws(ws6, "8E44AD", 95)

r6 = banner(ws6, 1, "SCREENSHOT ASSIGNMENT", "15 screenshots required  |  Sahil compiles all for final submission")
row_h(ws6, r6, 6); r6 += 1
r6 = thead(ws6, r6, ["#", "Description", "Owner", "Source Script", "When", "Filename", "Status"])

for i, s in enumerate(screenshots):
    r6 = trow(ws6, r6, list(s) + ["Not Started"], stripe=(i % 2 == 1), centers={0, 6})
    cell(ws6, r6 - 1, SC + 6, "Not Started", size=9, color=TEXT_MID, bg=GRAY_BG, h="center")

col_w(ws6, W_SHOT)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 - MEMBER QUICK REF
# ═══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Member Quick Ref")
setup_ws(ws7, NAVY, 95)

r7 = banner(ws7, 1, "MEMBER QUICK REFERENCE", "What you own | What you need first | Who to contact")
row_h(ws7, r7, 6); r7 += 1
r7 = thead(ws7, r7, ["Member", "Lab Task(s)", "Your Scripts", "Depends On", "Screenshots", "If Blocked Contact"])

for i, q in enumerate(quick_ref):
    r7 = trow(ws7, r7, list(q), stripe=(i % 2 == 1), bolds={0})

col_w(ws7, W_QUICK)

# ── Save ──────────────────────────────────────────────────────────────────────
out = r"c:\Users\Admin\Downloads\New folder (2)\Lab5\TEAM_WORK_ALLOCATION_NEW.xlsx"
alt = r"c:\Users\Admin\Downloads\New folder (2)\Lab5\TEAM_WORK_ALLOCATION.xlsx"
try:
    if os.path.exists(out):
        os.remove(out)
    wb.save(out)
except PermissionError:
    wb.save(alt)
    out = alt

print(f"Saved: {out}")
